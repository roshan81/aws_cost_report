import boto3
from datetime import datetime, date, timedelta
import os
import decimal
from botocore.exceptions import ClientError
import logging
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.cell import Cell

from email import encoders
from email.mime.base import MIMEBase
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

ASSUME_ROLE = os.environ['ASSUME_ROLE']

SES_REGION = os.environ['SES_REGION']

SUBJECT = "[AWS COSTING]: AWS Costs from {} to {}"

CHARSET = 'UTF-8'

SEND_FROM = os.environ['SEND_FROM']

ri_services = ['Amazon Elastic Compute Cloud - Compute', 'Amazon Relational Database Service', 'Amazon Redshift', 'Amazon ElastiCache', 'Amazon Elasticsearch Service', 'Amazon OpenSearch Service']

recipients = [
# add recepient emails to this list
    ]
    
aws_accounts = {
'123456789111':[{
    'acc_name':'aws-account-name01',
    'ano_arn':'arn:aws:ce::123456789111:anomalymonitor/84f93920-r8e90-333f-9320d-48wehr30',
    'acc_owner':'account.owner@email.com'
}],
'123456789011':[{
    'acc_name':'aws-account-name02',
    'ano_arn':'arn:aws:ce::123456789011:anomalymonitor/84f93920-480dfr-292f-3020f-48wehr30',
    'acc_owner':'account.owner@email.com'
}]
}


def get_client(instance_account_id: str, instance_region_id: str, resource_type: str):

    # Go for STS to assume role for cross account
    sts_connection = boto3.client('sts')
    acct_b = sts_connection.assume_role(
        RoleArn="arn:aws:iam::{}:role/{}".format(instance_account_id, ASSUME_ROLE),
        RoleSessionName="cross_acct_access_for_lambda"
    )
    access_key = acct_b['Credentials']['AccessKeyId']
    secret_key = acct_b['Credentials']['SecretAccessKey']
    session_token = acct_b['Credentials']['SessionToken']
     # create service client using the assumed role credentials
    return boto3.client(
        resource_type,
        aws_access_key_id=access_key,
        aws_secret_access_key=secret_key,
        aws_session_token=session_token,
        region_name=instance_region_id
    )

    
def get_week_days(year, week):

    year_start = date(year, 1, 1)

    week_start = year_start + timedelta(days=-year_start.isoweekday(), weeks=week)
    week_end_title = week_start + timedelta(days=6)
    week_end = week_start + timedelta(days=7)
    month_start = week_start.replace(day=1)
    return week_start, week_end, week_end_title, month_start


def upload_file(file_name, bucket, object_name=None):

    # If S3 object_name was not specified, use file_name
    if object_name is None:
        object_name = os.path.basename(file_name)

    s3_client = boto3.client('s3')
    try:
        response = s3_client.upload_file(file_name, bucket, object_name)
    except ClientError as e:
        logging.error(e)
        return False
    return True

def send_email_with_attachment(start_week, end_week_title, bodyhtml, file_name, ses_region):
    msg = MIMEMultipart()
    msg["Subject"] = SUBJECT.format(start_week, end_week_title)
    msg["From"] = SEND_FROM
    # convert recepients list to string
    msg["To"] = ", ".join(recipients)
    
    # Set message body
    body = MIMEText(bodyhtml.format(start_week, end_week_title), "html")
    msg.attach(body)

    for filename in file_name:

        with open(filename, "rb") as attachment:
            part = MIMEApplication(attachment.read())
            part.add_header("Content-Disposition",
                            "attachment",
                            filename=os.path.basename(filename))
        msg.attach(part)
    # Convert message to string and send
    ses_client = boto3.client("ses", region_name=ses_region)
    response = ses_client.send_raw_email(
        Source=SEND_FROM,
        Destinations=recipients,
        RawMessage={"Data": msg.as_string()}
    )
    print('Email sent to ', recipients, ' Message ID: ', response['MessageId'])



def lambda_handler(event, context):
    
    BODY_HTML = """<html>
        <head></head>
        <body>
        Hi,<br>
        As part of Cloud Governance, we are sending weekly AWS costing report along with Cost Anomalies and RI recommendations for each of the AWS accounts. The report is generated using the amortized costs of the account. Respective account owners can find any cost anomalies and recommendations of the accounts. There are two excel files attached for the top 10 spends and RI recommendations for each of the AWS accounts.<br>
        <br>The 'Total cost for all services this week' will be yellow highlighted if the allocated budget is exceeded for the week.<br>
        <br>This Report will shared once every week, please let us know if any of the recipients needed to be changed or added to this chain. 
        <br>
        <br>
        <h4>AWS costs for the period of {} to {}</h4>
        <h4>######################################</h4>
        <div><table style="width:100%">"""
    
    local_client = boto3.client('ce')


    # Finding start and end date of last week
    current_date = datetime.now()
    current_year = current_date.year
    last_week = int((current_date.strftime("%W")))-1
    week_days = get_week_days(current_year, last_week)

    start_week = week_days[0] 
    end_week = week_days[1]
    end_week_title = week_days[2]
    month_start_date = week_days[3]
    
    
    accounts = aws_accounts.keys()
    
    # create xlsx file attachments
    wb = Workbook()
    wb_ri = Workbook()
    bold_font = Font(bold=True)
    xlheaders = ['Account_ID', 'Account_Name', 'Service_Name', 'AWS_Cost','Anomaly_Service', 'Anomaly_StartDate', 'Anomaly_EndDate', 'Region', 'Usage_Type','Max_Anomaly_Impact','Total_Anomaly_Impact']
    worksheet_owner = set()
    #getting and creating sheet names from the dictionary
    for i in aws_accounts.values():
        for subdict in i:
            owner_email = subdict['acc_owner']
            owner_name = owner_email.split('@')[0]
            if owner_name not in worksheet_owner:
                worksheet_owner.add(owner_name)
                wb.create_sheet(owner_name)
                wb_ri.create_sheet(owner_name)

    

    #deleting unwanted initial sheet
    del wb['Sheet']
    del wb_ri['Sheet']

    # create column headers for each sheet
    for sheet in wb:
        sheet.append(xlheaders)
        for col in range(1, 12):
            sheet[get_column_letter(col) + '1'].font = Font(bold=True)
    

    for account in accounts:
        db = {}
        total_cost = 0
        if account == '123456789111':
            ##### Getting Cost and Usage data #####
            results = local_client.get_cost_and_usage(
                TimePeriod={
                    'Start': str(start_week),
                    'End': str(end_week)
                },
                Granularity='MONTHLY',
                Metrics=[
                    'AmortizedCost',
                ],
                GroupBy=[
                    {
                        'Type': 'DIMENSION',
                        'Key': 'SERVICE'
                    },
                ],
            )
            ##### Getting up to date cost
            month_results = local_client.get_cost_and_usage(
                TimePeriod={
                    'Start': str(month_start_date),
                    'End': str(end_week)
                },
                Granularity='MONTHLY',
                Metrics=[
                    'AmortizedCost',
                ]
            )
            if month_results.get('ResultsByTime'):
                uptodate_cost = month_results.get('ResultsByTime')[0]['Total']['AmortizedCost']['Amount']
                decimalMonth = decimal.Decimal(uptodate_cost)
                roundedMonth = decimalMonth.quantize(decimal.Decimal('0.00'))
            #print ("This is the results: {}".format(results))
            services = results.get('ResultsByTime')[0]['Groups']
            for service in services:
                service_name = service.get('Keys')[0]
                amount = service.get('Metrics')['AmortizedCost']['Amount']
                decimalAmount = decimal.Decimal(amount)
                roundedAmount = decimalAmount.quantize(decimal.Decimal('0.00'))
                db[service_name] = roundedAmount
                total_cost = total_cost + roundedAmount
                #BODY_HTML = BODY_HTML + "<tr><td>"+account+"</td><td>"+service_name+"</td><td>"+str(roundedAmount)+"</td></tr>"
            sorted_dict = sorted(db.items(), key=lambda x:x[1], reverse=True)
            sorted_dict = dict(sorted_dict)
            index = 0
            n = 1
            #print("Sorted Dictionary: {}".format(sorted_dict))
            BODY_HTML = BODY_HTML + "<h4>Account Owner: {}</h4>".format(aws_accounts[account][0]['acc_owner'])
            BODY_HTML = BODY_HTML + """<div><table style="width:100%">
            <thead><tr><td><b style="font-family:'Open Sans';font-size:13px">Account No</b></td><td><b style="font-family:'Open Sans';font-size:13px">Account Name</b></td><td><b style="font-family:'Open Sans';font-size:13px">Service Name</b></td><td><b style="font-family:'Open Sans';font-size:13px">AWS Cost</b></td></tr></thead>
            <tbody>"""
            for i in sorted_dict:
                index += 1
                if n == 1:
                    BODY_HTML = BODY_HTML + "<tr><td style=font-weight:bold>"+account+"</td><td style=font-weight:bold>"+aws_accounts[account][0]['acc_name']+"</td><td style=font-weight:bold>"+i+"</td><td style=font-weight:bold>"+str(sorted_dict[i])+"</td></tr>"
                    n += 1
                    #writing to excel
                    owner_name = aws_accounts[account][0]['acc_owner'].split('@')[0]
                    ws = wb[owner_name]
                    ws.append([
                        account,
                        aws_accounts[account][0]['acc_name'],
                        i,
                        str(sorted_dict[i]),
                        '-',
                        '-',
                        '-',
                        '-',
                        '-',
                        '-',
                        '-'
                    ])
                else:
                    BODY_HTML = BODY_HTML + "<tr><td>"+account+"</td><td>"+aws_accounts[account][0]['acc_name']+"</td><td>"+i+"</td><td>"+str(sorted_dict[i])+"</td></tr>"
                    #writing to excel
                    owner_name = aws_accounts[account][0]['acc_owner'].split('@')[0]
                    ws = wb[owner_name]
                    ws.append([
                        account,
                        aws_accounts[account][0]['acc_name'],
                        i,
                        str(sorted_dict[i]),
                        '-',
                        '-',
                        '-',
                        '-',
                        '-',
                        '-',
                        '-'
                    ])                        
                    if index > 9:
                        break
                    
            ##### Getting monthly budget limit #####
            budget_client = boto3.client('budgets')
            budget_name = 'finops-ACCOUNT_' + account + '_MONTHLY_' + str(current_year)
            try:
                budget_response = budget_client.describe_budget(
                    AccountId = account,
                    BudgetName = budget_name
                    )
                if budget_response.get('Budget'):
                    budgeted_amount = budget_response['Budget']['BudgetLimit'].get('Amount')
                    weekly_budget = decimal.Decimal(budgeted_amount) / 4
                    weekly_budget_rounded = weekly_budget.quantize(decimal.Decimal('0.00'))
                    if total_cost > weekly_budget_rounded:
                        BODY_HTML = BODY_HTML + "<tr height = 20px></tr>"
                        BODY_HTML = BODY_HTML + "<tr colspan=2><td><span style='font-weight:bold;background-color: yellow'>"+"[!!Weekly budget limit exceeded!!] Total cost for all services this week == ${}".format(total_cost)+"</span></td>"
                    else:
                        BODY_HTML = BODY_HTML + "<tr height = 20px></tr>"
                        BODY_HTML = BODY_HTML + "<tr colspan=2><td><span style=font-weight:bold>"+"Total cost for all services this week == ${}".format(total_cost)+"</span></td>"
                
            except ClientError:
                BODY_HTML = BODY_HTML + "<tr colspan=2><td><span style=font-weight:bold>"+"Total cost for all services this week == ${}".format(total_cost)+"</span></td>"
            BODY_HTML = BODY_HTML + "<tr colspan=2><td><span style=font-weight:bold>"+"Overall spend for the account in this month == ${}".format(roundedMonth)+"</span></td>"
            BODY_HTML = BODY_HTML + "<tr height = 20px></tr>"
            BODY_HTML = BODY_HTML + "</tbody></table></div>"
            ##### Getting Anomaly data #####
            ano_results = local_client.get_anomalies(
                DateInterval={
                    'StartDate': str(start_week),
                    'EndDate': str(end_week)
                },
                MonitorArn = aws_accounts[account][0]['ano_arn']
            )
            if ano_results['Anomalies']:
                BODY_HTML = BODY_HTML + "<h4>Anomaly Details for the account {0} ({1}):</h4>".format(account, aws_accounts[account][0]['acc_name'])
                BODY_HTML = BODY_HTML + """<div><table style="width:100%">
                <thead><tr><td><b style="font-family:'Open Sans';font-size:13px">Service</b></td><td><b style="font-family:'Open Sans';font-size:13px">Start Date</b></td><td><b style="font-family:'Open Sans';font-size:13px">End Date</b></td><td><b style="font-family:'Open Sans';font-size:13px">Region</b></td><td><b style="font-family:'Open Sans';font-size:13px">UsageType</b></td><td><b style="font-family:'Open Sans';font-size:13px">Max Impact</b></td><td><b style="font-family:'Open Sans';font-size:13px">Total Impact</b></td></tr></thead>
                <tbody>"""
                details = ano_results['Anomalies']
                #print("Details: {}".format(details))
                maximpact = {}
                index = 0
                for element in details:
                    maximpact[index] = element['Impact']['MaxImpact']
                    index += 1
                #print("MAx impact: {}".format(maximpact))
                sorted_max = sorted(maximpact.items(), key=lambda x:x[1], reverse=True)
                sorted_max = dict(sorted_max)
                #print("Sorted maximpact: {}".format(sorted_max))
                #print("Sortex max: {}".format(sorted_max.keys()))
                for i in sorted_max.keys():
                    detail = details[i]
                    #print("Detail: {}".format(detail))
                    ano_startdate = ((detail['AnomalyStartDate']).split("T"))[0]
                    ano_enddate = ((detail['AnomalyEndDate']).split("T"))[0]
                    if detail['RootCauses'][0].get('Region'):
                        region = detail['RootCauses'][0]['Region']
                    else:
                        region = '-'
                    if detail['RootCauses'][0].get('UsageType'):
                        usage_type = detail['RootCauses'][0]['UsageType']
                    else:
                        usage_type = '-'
                    BODY_HTML = BODY_HTML + "<tr><td>"+detail['RootCauses'][0]['Service']+"</td><td>"+ano_startdate+"</td><td>"+ano_enddate+"</td><td>"+region+"</td><td>"+usage_type+"</td><td>"+str(detail['Impact']['MaxImpact'])+"</td><td>"+str(detail['Impact']['TotalImpact'])+"</td></tr>"
                    #writing to excel
                    owner_name = aws_accounts[account][0]['acc_owner'].split('@')[0]
                    ws = wb[owner_name]
                    ws.append([
                        account,
                        aws_accounts[account][0]['acc_name'],
                        '-',
                        '-',
                        detail['RootCauses'][0]['Service'],
                        ano_startdate,
                        ano_enddate,
                        region,
                        usage_type,
                        str(detail['Impact']['MaxImpact']),
                        str(detail['Impact']['TotalImpact'])
                    ])                                        
                BODY_HTML = BODY_HTML + "<tr height = 20px></tr>"
                BODY_HTML = BODY_HTML + "</tbody></table></div>"
                
            ##### Getting RI recommendations #####
            for riservice in ri_services:
                ri_results = local_client.get_reservation_purchase_recommendation(
                    AccountId = account,
                    Service = riservice,
                    TermInYears = 'ONE_YEAR',
                    LookbackPeriodInDays = 'THIRTY_DAYS',
                    PaymentOption = 'ALL_UPFRONT'
                )
                
                if ri_results.get('Recommendations'):
                    #BODY_HTML = BODY_HTML + "<h4>RI Recommendations for the account {0} ({1}):</h4>".format(account, aws_accounts[account][0]['acc_name'])
                    
                    if riservice == 'Amazon Relational Database Service':
                        BODY_HTML = BODY_HTML + "<h4>RI Recommendations for {0} (RDS):</h4>".format(riservice)
                        BODY_HTML = BODY_HTML + """<div><table style="width:100%">
                        <thead><tr><td><b style="font-family:'Open Sans';font-size:13px">Action</b></td><td><b style="font-family:'Open Sans';font-size:13px">Instance Type</b></td><td><b style="font-family:'Open Sans';font-size:13px">Region</b></td><td><b style="font-family:'Open Sans';font-size:13px">Database</b></td><td><b style="font-family:'Open Sans';font-size:13px">License</b></td><td><b style="font-family:'Open Sans';font-size:13px">Current Generation</b></td><td><b style="font-family:'Open Sans';font-size:13px">Upfront Cost</b></td><td><b style="font-family:'Open Sans';font-size:13px">Estimated Monthly Savings</b></td></tr></thead>
                        <tbody>"""                        
                        ri_details = ri_results['Recommendations'][0]['RecommendationDetails']
                        maximpact = {}
                        index = 0
                        for ri in ri_details:
                            maximpact[index] = decimal.Decimal(ri['EstimatedMonthlySavingsAmount'])
                            index += 1
                        sort_max = sorted(maximpact.items(), key=lambda x:x[1], reverse=True)
                        sort_max = dict(sort_max)

                        # writing to excel
                        owner_name = aws_accounts[account][0]['acc_owner'].split('@')[0]
                        ws = wb_ri[owner_name]
                        ws.append([])
                        bold_cell = Cell(ws, value='Amazon RDS for account {}'.format(aws_accounts[account][0]['acc_name']))
                        ws.append([bold_cell])
                        ws.append(['Action', 'Instance Type', 'Region', 'Database', 'License', 'Current Generation', 'Upfront Cost', 'Estimated Monthly Savings'])

                        for i in sort_max.keys():
                            detail = ri_details[i]
                            action = "Buy {0} {1}".format(detail['RecommendedNumberOfInstancesToPurchase'],detail['InstanceDetails']['RDSInstanceDetails']['InstanceType'])
                            instancetype = detail['InstanceDetails']['RDSInstanceDetails']['InstanceType']
                            rgn = detail['InstanceDetails']['RDSInstanceDetails']['Region']
                            database = detail['InstanceDetails']['RDSInstanceDetails']['DatabaseEngine']
                            licen = detail['InstanceDetails']['RDSInstanceDetails']['LicenseModel']
                            current_gen = str(detail['InstanceDetails']['RDSInstanceDetails']['CurrentGeneration'])
                            upcost = decimal.Decimal(detail['UpfrontCost'])
                            upcost_rounded = str(upcost.quantize(decimal.Decimal('0.00')))
                            est_saving = decimal.Decimal(detail['EstimatedMonthlySavingsAmount'])
                            est_sav_rounded = str(est_saving.quantize(decimal.Decimal('0.00')))
                            BODY_HTML = BODY_HTML + "<tr><td>"+action+"</td><td>"+instancetype+"</td><td>"+rgn+"</td><td>"+database+"</td><td>"+licen+"</td><td>"+current_gen+"</td><td>"+upcost_rounded+"</td><td>"+est_sav_rounded+"</td></tr>"
                            # writing to excel
                            ws.append([
                                action,
                                instancetype,
                                rgn,
                                database,
                                licen,
                                current_gen,
                                upcost_rounded,
                                est_sav_rounded
                                ])
                            

                        BODY_HTML = BODY_HTML + "<tr height = 20px></tr>"
                        BODY_HTML = BODY_HTML + "</tbody></table></div>"
                    
                    elif riservice == 'Amazon ElastiCache':
                        BODY_HTML = BODY_HTML + "<h4>RI Recommendations for {0}:</h4>".format(riservice)
                        BODY_HTML = BODY_HTML + """<div><table style="width:100%">
                        <thead><tr><td><b style="font-family:'Open Sans';font-size:13px">Action</b></td><td><b style="font-family:'Open Sans';font-size:13px">Instance Type</b></td><td><b style="font-family:'Open Sans';font-size:13px">Region</b></td><td><b style="font-family:'Open Sans';font-size:13px">Cache Engine</b></td><td><b style="font-family:'Open Sans';font-size:13px">Current Generation</b></td><td><b style="font-family:'Open Sans';font-size:13px">Upfront Cost</b></td><td><b style="font-family:'Open Sans';font-size:13px">Estimated Monthly Savings</b></td></tr></thead>
                        <tbody>"""
                        ri_details = ri_results['Recommendations'][0]['RecommendationDetails']
                        max_impact = {}
                        index = 0
                        
                        for ri in ri_details:
                            max_impact[index] = decimal.Decimal(ri['EstimatedMonthlySavingsAmount'])
                            index += 1
                        sort_max = sorted(max_impact.items(), key=lambda x:x[1], reverse=True)
                        sort_max = dict(sort_max)

                        # writing to excel
                        owner_name = aws_accounts[account][0]['acc_owner'].split('@')[0]
                        ws = wb_ri[owner_name]
                        ws.append([])
                        bold_cell = Cell(ws, value='Amazon ElastiCache for account {}'.format(aws_accounts[account][0]['acc_name']))
                        ws.append([bold_cell])
                        ws.append(['Action', 'Instance Type', 'Region', 'Cache Engine', 'Current Generation', 'Upfront Cost', 'Estimated Monthly Savings'])

                        for i in sort_max:
                            detail = ri_details[i]
                            
                            action = "Buy {0} {1}".format(detail['RecommendedNumberOfInstancesToPurchase'],detail['InstanceDetails']['ElastiCacheInstanceDetails']['NodeType'])
                            instancetype = detail['InstanceDetails']['ElastiCacheInstanceDetails']['NodeType']
                            rgn = detail['InstanceDetails']['ElastiCacheInstanceDetails']['Region']
                            engine = detail['InstanceDetails']['ElastiCacheInstanceDetails']['ProductDescription']
                            current_gen = str(detail['InstanceDetails']['ElastiCacheInstanceDetails']['CurrentGeneration'])
                            upcost = decimal.Decimal(detail['UpfrontCost'])
                            upcost_rounded = str(upcost.quantize(decimal.Decimal('0.00')))
                            est_saving = decimal.Decimal(detail['EstimatedMonthlySavingsAmount'])
                            est_sav_rounded = str(est_saving.quantize(decimal.Decimal('0.00')))
                            BODY_HTML = BODY_HTML + "<tr><td>"+action+"</td><td>"+instancetype+"</td><td>"+rgn+"</td><td>"+engine+"</td><td>"+current_gen+"</td><td>"+upcost_rounded+"</td><td>"+est_sav_rounded+"</td></tr>"
                            # writing to excel
                            ws.append([
                                action,
                                instancetype,
                                rgn,
                                engine,
                                current_gen,
                                upcost_rounded,
                                est_sav_rounded
                            ])
                        BODY_HTML = BODY_HTML + "<tr height = 20px></tr>"
                        BODY_HTML = BODY_HTML + "</tbody></table></div>"
                        
                    elif riservice == 'Amazon OpenSearch Service':
                        BODY_HTML = BODY_HTML + "<h4>RI Recommendations for {0}:</h4>".format(riservice)
                        BODY_HTML = BODY_HTML + """<div><table style="width:100%">
                        <thead><tr><td><b style="font-family:'Open Sans';font-size:13px">Action</b></td><td><b style="font-family:'Open Sans';font-size:13px">Instance Type</b></td><td><b style="font-family:'Open Sans';font-size:13px">Region</b></td><td><b style="font-family:'Open Sans';font-size:13px">Current Generation</b></td><td><b style="font-family:'Open Sans';font-size:13px">Upfront Cost</b></td><td><b style="font-family:'Open Sans';font-size:13px">Estimated Monthly Savings</b></td></tr></thead>
                        <tbody>"""
                        ri_details = ri_results['Recommendations'][0]['RecommendationDetails']
                        max_impact = {}
                        index = 0
                        
                        for ri in ri_details:
                            max_impact[index] = decimal.Decimal(ri['EstimatedMonthlySavingsAmount'])
                            index += 1
                        sort_max = sorted(max_impact.items(), key=lambda x:x[1], reverse=True)
                        sort_max = dict(sort_max)

                        # writing to excel
                        owner_name = aws_accounts[account][0]['acc_owner'].split('@')[0]
                        ws = wb_ri[owner_name]
                        ws.append([])
                        bold_cell = Cell(ws, value='Amazon OpenSearch Service for account {}'.format(aws_accounts[account][0]['acc_name']))
                        ws.append([bold_cell])
                        ws.append(['Action', 'Instance Type', 'Region', 'Current Generation', 'Upfront Cost', 'Estimated Monthly Savings'])

                        for i in sort_max:
                            detail = ri_details[i]
                            
                            action = "Buy {0} {1}".format(detail['RecommendedNumberOfInstancesToPurchase'],detail['InstanceDetails']['ESInstanceDetails']['InstanceSize'])
                            instancetype = detail['InstanceDetails']['ESInstanceDetails']['InstanceSize']
                            rgn = detail['InstanceDetails']['ESInstanceDetails']['Region']
                            current_gen = str(detail['InstanceDetails']['ESInstanceDetails']['CurrentGeneration'])
                            upcost = decimal.Decimal(detail['UpfrontCost'])
                            upcost_rounded = str(upcost.quantize(decimal.Decimal('0.00')))
                            est_saving = decimal.Decimal(detail['EstimatedMonthlySavingsAmount'])
                            est_sav_rounded = str(est_saving.quantize(decimal.Decimal('0.00')))
                            BODY_HTML = BODY_HTML + "<tr><td>"+action+"</td><td>"+instancetype+"</td><td>"+rgn+"</td><td>"+current_gen+"</td><td>"+upcost_rounded+"</td><td>"+est_sav_rounded+"</td></tr>"
                            # writing to excel
                            ws.append([
                                action,
                                instancetype,
                                rgn,
                                current_gen,
                                upcost_rounded,
                                est_sav_rounded
                            ])
                        BODY_HTML = BODY_HTML + "<tr height = 20px></tr>"
                        BODY_HTML = BODY_HTML + "</tbody></table></div>"
                    
                    elif riservice == 'Amazon Elastic Compute Cloud - Compute':
                        BODY_HTML = BODY_HTML + "<h4>RI Recommendations for {0}:</h4>".format(riservice)
                        BODY_HTML = BODY_HTML + """<div><table style="width:100%">
                        <thead><tr><td><b style="font-family:'Open Sans';font-size:13px">Action</b></td><td><b style="font-family:'Open Sans';font-size:13px">Instance Type</b></td><td><b style="font-family:'Open Sans';font-size:13px">Platform</b></td><td><b style="font-family:'Open Sans';font-size:13px">Region</b></td><td><b style="font-family:'Open Sans';font-size:13px">Current Generation</b></td><td><b style="font-family:'Open Sans';font-size:13px">Upfront Cost</b></td><td><b style="font-family:'Open Sans';font-size:13px">Estimated Monthly Savings</b></td></tr></thead>
                        <tbody>"""
                        ri_details = ri_results['Recommendations'][0]['RecommendationDetails']
                        max_impact = {}
                        index = 0
                        
                        for ri in ri_details:
                            max_impact[index] = decimal.Decimal(ri['EstimatedMonthlySavingsAmount'])
                            index += 1
                        sort_max = sorted(max_impact.items(), key=lambda x:x[1], reverse=True)
                        sort_max = dict(sort_max)

                        # writing to excel
                        owner_name = aws_accounts[account][0]['acc_owner'].split('@')[0]
                        ws = wb_ri[owner_name]
                        ws.append([])
                        bold_cell = Cell(ws, value='Amazon Elastic Compute Cloud - Compute for account {}'.format(aws_accounts[account][0]['acc_name']))
                        ws.append([bold_cell])
                        ws.append(['Action', 'Instance Type', 'Platform', 'Region', 'Current Generation', 'Upfront Cost', 'Estimated Monthly Savings'])

                        for i in sort_max:
                            detail = ri_details[i]
                            
                            action = "Buy {0} {1}".format(detail['RecommendedNumberOfInstancesToPurchase'],detail['InstanceDetails']['EC2InstanceDetails']['InstanceType'])
                            instancetype = detail['InstanceDetails']['EC2InstanceDetails']['InstanceType']
                            platform = detail['InstanceDetails']['EC2InstanceDetails']['Platform']
                            rgn = detail['InstanceDetails']['EC2InstanceDetails']['Region']
                            current_gen = str(detail['InstanceDetails']['EC2InstanceDetails']['CurrentGeneration'])
                            upcost = decimal.Decimal(detail['UpfrontCost'])
                            upcost_rounded = str(upcost.quantize(decimal.Decimal('0.00')))
                            est_saving = decimal.Decimal(detail['EstimatedMonthlySavingsAmount'])
                            est_sav_rounded = str(est_saving.quantize(decimal.Decimal('0.00')))
                            BODY_HTML = BODY_HTML + "<tr><td>"+action+"</td><td>"+instancetype+"</td><td>"+platform+"</td><td>"+rgn+"</td><td>"+current_gen+"</td><td>"+upcost_rounded+"</td><td>"+est_sav_rounded+"</td></tr>"
                            # writing to excel
                            ws.append([
                                action,
                                instancetype,
                                platform,
                                rgn,
                                current_gen,
                                upcost_rounded,
                                est_sav_rounded
                            ])
                        BODY_HTML = BODY_HTML + "<tr height = 20px></tr>"
                        BODY_HTML = BODY_HTML + "</tbody></table></div>"
                    elif riservice == 'Amazon Redshift':
                        BODY_HTML = BODY_HTML + "<h4>RI Recommendations for {0}:</h4>".format(riservice)
                        BODY_HTML = BODY_HTML + """<div><table style="width:100%">
                        <thead><tr><td><b style="font-family:'Open Sans';font-size:13px">Action</b></td><td><b style="font-family:'Open Sans';font-size:13px">Instance Type</b></td><td><b style="font-family:'Open Sans';font-size:13px">Region</b></td><td><b style="font-family:'Open Sans';font-size:13px">SizeFlex Eligible</b></td><td><b style="font-family:'Open Sans';font-size:13px">Current Generation</b></td><td><b style="font-family:'Open Sans';font-size:13px">Upfront Cost</b></td><td><b style="font-family:'Open Sans';font-size:13px">Estimated Monthly Savings</b></td></tr></thead>
                        <tbody>"""
                        ri_details = ri_results['Recommendations'][0]['RecommendationDetails']
                        max_impact = {}
                        index = 0
                        
                        for ri in ri_details:
                            max_impact[index] = decimal.Decimal(ri['EstimatedMonthlySavingsAmount'])
                            index += 1
                        sort_max = sorted(max_impact.items(), key=lambda x:x[1], reverse=True)
                        sort_max = dict(sort_max)

                        # writing to excel
                        owner_name = aws_accounts[account][0]['acc_owner'].split('@')[0]
                        ws = wb_ri[owner_name]
                        ws.append([])
                        bold_cell = Cell(ws, value='Amazon Redshift for account {}'.format(aws_accounts[account][0]['acc_name']))
                        ws.append([bold_cell])
                        ws.append(['Action', 'Instance Type', 'Region', 'SizeFlex Eligible', 'Current Generation', 'Upfront Cost', 'Estimated Monthly Savings'])

                        for i in sort_max:
                            detail = ri_details[i]
                            
                            action = "Buy {0} {1}".format(detail['RecommendedNumberOfInstancesToPurchase'],detail['InstanceDetails']['RedshiftInstanceDetails']['NodeType'])
                            instancetype = detail['InstanceDetails']['RedshiftInstanceDetails']['NodeType']
                            rgn = detail['InstanceDetails']['RedshiftInstanceDetails']['Region']
                            sizeflex = str(detail['InstanceDetails']['RedshiftInstanceDetails']['SizeFlexEligible'])
                            current_gen = str(detail['InstanceDetails']['RedshiftInstanceDetails']['CurrentGeneration'])
                            upcost = decimal.Decimal(detail['UpfrontCost'])
                            upcost_rounded = str(upcost.quantize(decimal.Decimal('0.00')))
                            est_saving = decimal.Decimal(detail['EstimatedMonthlySavingsAmount'])
                            est_sav_rounded = str(est_saving.quantize(decimal.Decimal('0.00')))
                            BODY_HTML = BODY_HTML + "<tr><td>"+action+"</td><td>"+instancetype+"</td><td>"+rgn+"</td><td>"+sizeflex+"</td><td>"+current_gen+"</td><td>"+upcost_rounded+"</td><td>"+est_sav_rounded+"</td></tr>"
                            # writing to excel
                            ws.append([
                                action,
                                instancetype,
                                rgn,
                                sizeflex,
                                current_gen,
                                upcost_rounded,
                                est_sav_rounded
                            ])
                        BODY_HTML = BODY_HTML + "<tr height = 20px></tr>"
                        BODY_HTML = BODY_HTML + "</tbody></table></div>"
                        
            BODY_HTML = BODY_HTML + "<h4>######################################</h4>"
        else:
            remote_client = get_client(account, 'us-east-1', 'ce')
            results = remote_client.get_cost_and_usage(
                TimePeriod={
                    'Start': str(start_week),
                    'End': str(end_week)
                },
                Granularity='MONTHLY',
                Metrics=[
                    'AmortizedCost',
                ],
                GroupBy=[
                    {
                        'Type': 'DIMENSION',
                        'Key': 'SERVICE'
                    },
                ],
            )
            month_results = remote_client.get_cost_and_usage(
                TimePeriod={
                    'Start': str(month_start_date),
                    'End': str(end_week)
                },
                Granularity='MONTHLY',
                Metrics=[
                    'AmortizedCost',
                ]
            )
            if month_results.get('ResultsByTime'):
                uptodate_cost = month_results.get('ResultsByTime')[0]['Total']['AmortizedCost']['Amount']
                decimalMonth = decimal.Decimal(uptodate_cost)
                roundedMonth = decimalMonth.quantize(decimal.Decimal('0.00'))
            get_services = results.get('ResultsByTime')[0]['Groups']
            for service in get_services:
                service_name = service.get('Keys')[0]
                amount = service.get('Metrics')['AmortizedCost']['Amount']
                decimalAmount = decimal.Decimal(amount)
                roundedAmount = decimalAmount.quantize(decimal.Decimal('0.00'))
                db[service_name] = roundedAmount
                total_cost = total_cost + roundedAmount
                #BODY_HTML = BODY_HTML + "<tr><td>"+account+"</td><td>"+service_name+"</td><td>"+str(roundedAmount)+"</td></tr>"
            sorted_dict = sorted(db.items(), key=lambda x:x[1], reverse=True)
            sorted_dict = dict(sorted_dict)
            index = 0
            n = 1
            #print("Sorted Dictionary: {}".format(sorted_dict))
            BODY_HTML = BODY_HTML + "<h4>Account Owner: {}</h4>".format(aws_accounts[account][0]['acc_owner'])
            BODY_HTML = BODY_HTML + """<div><table style="width:100%">
            <thead><tr><td><b style="font-family:'Open Sans';font-size:13px">Account No</b></td><td><b style="font-family:'Open Sans';font-size:13px">Account Name</b></td><td><b style="font-family:'Open Sans';font-size:13px">Service Name</b></td><td><b style="font-family:'Open Sans';font-size:13px">AWS Cost</b></td></tr></thead>
            <tbody>"""
            for i in sorted_dict:
                index += 1
                if n == 1:
                    BODY_HTML = BODY_HTML + "<tr><td style=font-weight:bold>"+account+"</td><td style=font-weight:bold>"+aws_accounts[account][0]['acc_name']+"</td><td style=font-weight:bold>"+i+"</td><td style=font-weight:bold>"+str(sorted_dict[i])+"</td></tr>"
                    n += 1
                    #writing to excel
                    owner_name = aws_accounts[account][0]['acc_owner'].split('@')[0]
                    ws = wb[owner_name]
                    ws.append([
                        account,
                        aws_accounts[account][0]['acc_name'],
                        i,
                        str(sorted_dict[i]),
                        '-',
                        '-',
                        '-',
                        '-',
                        '-',
                        '-',
                        '-'
                    ])
                else:
                    BODY_HTML = BODY_HTML + "<tr><td>"+account+"</td><td>"+aws_accounts[account][0]['acc_name']+"</td><td>"+i+"</td><td>"+str(sorted_dict[i])+"</td></tr>"
                    #writing to excel
                    owner_name = aws_accounts[account][0]['acc_owner'].split('@')[0]
                    ws = wb[owner_name]
                    ws.append([
                        account,
                        aws_accounts[account][0]['acc_name'],
                        i,
                        str(sorted_dict[i]),
                        '-',
                        '-',
                        '-',
                        '-',
                        '-',
                        '-',
                        '-'
                    ])                
                    if index > 9:
                        break
                    
            ##### Getting monthly budget limit #####
            budget_client = get_client(account, 'us-east-1', 'budgets')
            budget_name = 'finops-ACCOUNT_' + account + '_MONTHLY_' + str(current_year)
            try:
                budget_response = budget_client.describe_budget(
                    AccountId = account,
                    BudgetName = budget_name
                    )
                if budget_response.get('Budget'):
                    budgeted_amount = budget_response['Budget']['BudgetLimit'].get('Amount')
                    weekly_budget = decimal.Decimal(budgeted_amount) / 4
                    weekly_budget_rounded = weekly_budget.quantize(decimal.Decimal('0.00'))
                    if total_cost > weekly_budget_rounded:
                        BODY_HTML = BODY_HTML + "<tr height = 20px></tr>"
                        BODY_HTML = BODY_HTML + "<tr colspan=2><td><span style='font-weight:bold;background-color: yellow'>"+"[!!Weekly budget limit exceeded!!] Total cost for all services this week == ${}".format(total_cost)+"</span></td>"
                    else:
                        BODY_HTML = BODY_HTML + "<tr height = 20px></tr>"
                        BODY_HTML = BODY_HTML + "<tr colspan=2><td><span style=font-weight:bold>"+"Total cost for all services this week == ${}".format(total_cost)+"</span></td>"
                
            except ClientError:
                budget_response = budget_client.describe_budgets(
                    AccountId = account
                )
                if budget_response.get('Budgets'):
                    budgeted_amount = budget_response['Budgets'][0]['BudgetLimit'].get('Amount')
                    weekly_budget = decimal.Decimal(budgeted_amount) / 4
                    weekly_budget_rounded = weekly_budget.quantize(decimal.Decimal('0.00'))
                    if total_cost > weekly_budget_rounded:
                        BODY_HTML = BODY_HTML + "<tr height = 20px></tr>"
                        BODY_HTML = BODY_HTML + "<tr colspan=2><td><span style='font-weight:bold;background-color: yellow'>"+"[!!Weekly budget limit exceeded!!] Total cost for all services this week == ${}".format(total_cost)+"</span></td>"
                    else:
                        BODY_HTML = BODY_HTML + "<tr height = 20px></tr>"
                        BODY_HTML = BODY_HTML + "<tr colspan=2><td><span style=font-weight:bold>"+"Total cost for all services this week == ${}".format(total_cost)+"</span></td>"
                #BODY_HTML = BODY_HTML + "<tr colspan=2><td><span style=font-weight:bold>"+"Total cost for all services this week == ${}".format(total_cost)+"</span></td>"
            BODY_HTML = BODY_HTML + "<tr colspan=2><td><span style=font-weight:bold>"+"Overall spend for the account in this month == ${}".format(roundedMonth)+"</span></td>"
            BODY_HTML = BODY_HTML + "<tr height = 20px></tr>"
            BODY_HTML = BODY_HTML + "</tbody></table></div>"
            ##### Getting Anomaly data #####
            ano_results = remote_client.get_anomalies(
                DateInterval={
                    'StartDate': str(start_week),
                    'EndDate': str(end_week)
                },
                MonitorArn = aws_accounts[account][0]['ano_arn']
            )
            if ano_results['Anomalies']:
                BODY_HTML = BODY_HTML + "<h4>Anomaly Details for the account {0} ({1}):</h4>".format(account, aws_accounts[account][0]['acc_name'])
                BODY_HTML = BODY_HTML + """<div><table style="width:100%">
                <thead><tr><td><b style="font-family:'Open Sans';font-size:13px">Service</b></td><td><b style="font-family:'Open Sans';font-size:13px">Start Date</b></td><td><b style="font-family:'Open Sans';font-size:13px">End Date</b></td><td><b style="font-family:'Open Sans';font-size:13px">Region</b></td><td><b style="font-family:'Open Sans';font-size:13px">UsageType</b></td><td><b style="font-family:'Open Sans';font-size:13px">Max Impact</b></td><td><b style="font-family:'Open Sans';font-size:13px">Total Impact</b></td></tr></thead>
                <tbody>"""
                details = ano_results['Anomalies']
                maximpact = {}
                index = 0
                for element in details:
                    maximpact[index] = element['Impact']['MaxImpact']
                    index += 1
                sort_max = sorted(maximpact.items(), key=lambda x:x[1], reverse=True)
                sort_max = dict(sort_max)
                for i in sort_max.keys():
                    detail = details[i]
                    ano_startdate = ((detail['AnomalyStartDate']).split("T"))[0]
                    ano_enddate = ((detail['AnomalyEndDate']).split("T"))[0]
                    if detail['RootCauses']:
                        if detail['RootCauses'][0].get('Region'):
                            region = detail['RootCauses'][0]['Region']
                        else:
                            region = '-'
                    else:
                        region = '-'
                        
                    if detail['RootCauses']:
                        if detail['RootCauses'][0].get('UsageType'):
                            usage_type = detail['RootCauses'][0]['UsageType']
                        else:
                            usage_type = '-'
                    else:
                        usage_type = '-'
                    BODY_HTML = BODY_HTML + "<tr><td>"+detail['DimensionValue']+"</td><td>"+ano_startdate+"</td><td>"+ano_enddate+"</td><td>"+region+"</td><td>"+usage_type+"</td><td>"+str(detail['Impact']['MaxImpact'])+"</td><td>"+str(detail['Impact']['TotalImpact'])+"</td></tr>"
                    #writing to excel
                    owner_name = aws_accounts[account][0]['acc_owner'].split('@')[0]
                    ws = wb[owner_name]
                    ws.append([
                        account,
                        aws_accounts[account][0]['acc_name'],
                        '-',
                        '-',
                        detail['DimensionValue'],
                        ano_startdate,
                        ano_enddate,
                        region,
                        usage_type,
                        str(detail['Impact']['MaxImpact']),
                        str(detail['Impact']['TotalImpact'])
                    ])                            
                BODY_HTML = BODY_HTML + "<tr height = 20px></tr>"
                BODY_HTML = BODY_HTML + "</tbody></table></div>"
            
            ##### Getting RI recommendations #####
            for riservice in ri_services:
                ri_results = remote_client.get_reservation_purchase_recommendation(
                    AccountId = account,
                    Service = riservice,
                    TermInYears = 'ONE_YEAR',
                    LookbackPeriodInDays = 'THIRTY_DAYS',
                    PaymentOption = 'ALL_UPFRONT'
                )
                
                if ri_results.get('Recommendations'):
                    #BODY_HTML = BODY_HTML + "<h4>RI Recommendations for the account {0} ({1}):</h4>".format(account, aws_accounts[account][0]['acc_name'])
                    
                    if riservice == 'Amazon Relational Database Service':
                        BODY_HTML = BODY_HTML + "<h4>RI Recommendations for {0} (RDS):</h4>".format(riservice)
                        BODY_HTML = BODY_HTML + """<div><table style="width:100%">
                        <thead><tr><td><b style="font-family:'Open Sans';font-size:13px">Action</b></td><td><b style="font-family:'Open Sans';font-size:13px">Instance Type</b></td><td><b style="font-family:'Open Sans';font-size:13px">Region</b></td><td><b style="font-family:'Open Sans';font-size:13px">Database</b></td><td><b style="font-family:'Open Sans';font-size:13px">License</b></td><td><b style="font-family:'Open Sans';font-size:13px">Current Generation</b></td><td><b style="font-family:'Open Sans';font-size:13px">Upfront Cost</b></td><td><b style="font-family:'Open Sans';font-size:13px">Estimated Monthly Savings</b></td></tr></thead>
                        <tbody>"""                        
                        ri_details = ri_results['Recommendations'][0]['RecommendationDetails']
                        maximpact = {}
                        index = 0
                        for ri in ri_details:
                            maximpact[index] = decimal.Decimal(ri['EstimatedMonthlySavingsAmount'])
                            index += 1
                        sort_max = sorted(maximpact.items(), key=lambda x:x[1], reverse=True)
                        sort_max = dict(sort_max)

                        # writing to excel
                        owner_name = aws_accounts[account][0]['acc_owner'].split('@')[0]
                        ws = wb_ri[owner_name]
                        ws.append([])
                        bold_cell = Cell(ws, value='Amazon RDS for account {}'.format(aws_accounts[account][0]['acc_name']))
                        ws.append([bold_cell])
                        ws.append(['Action', 'Instance Type', 'Region', 'Database', 'License', 'Current Generation', 'Upfront Cost', 'Estimated Monthly Savings'])
                        
                        for i in sort_max.keys():
                            detail = ri_details[i]
                            action = "Buy {0} {1}".format(detail['RecommendedNumberOfInstancesToPurchase'],detail['InstanceDetails']['RDSInstanceDetails']['InstanceType'])
                            instancetype = detail['InstanceDetails']['RDSInstanceDetails']['InstanceType']
                            rgn = detail['InstanceDetails']['RDSInstanceDetails']['Region']
                            database = detail['InstanceDetails']['RDSInstanceDetails']['DatabaseEngine']
                            licen = detail['InstanceDetails']['RDSInstanceDetails']['LicenseModel']
                            current_gen = str(detail['InstanceDetails']['RDSInstanceDetails']['CurrentGeneration'])
                            upcost = decimal.Decimal(detail['UpfrontCost'])
                            upcost_rounded = str(upcost.quantize(decimal.Decimal('0.00')))
                            est_saving = decimal.Decimal(detail['EstimatedMonthlySavingsAmount'])
                            est_sav_rounded = str(est_saving.quantize(decimal.Decimal('0.00')))
                            
                            BODY_HTML = BODY_HTML + "<tr><td>"+action+"</td><td>"+instancetype+"</td><td>"+rgn+"</td><td>"+database+"</td><td>"+licen+"</td><td>"+current_gen+"</td><td>"+upcost_rounded+"</td><td>"+est_sav_rounded+"</td></tr>"
                            # writing to excel
                            ws.append([
                                action,
                                instancetype,
                                rgn,
                                database,
                                licen,
                                current_gen,
                                upcost_rounded,
                                est_sav_rounded
                                ])

                        BODY_HTML = BODY_HTML + "<tr height = 20px></tr>"
                        BODY_HTML = BODY_HTML + "</tbody></table></div>"
                    
                    elif riservice == 'Amazon ElastiCache':
                        BODY_HTML = BODY_HTML + "<h4>RI Recommendations for {0}:</h4>".format(riservice)
                        BODY_HTML = BODY_HTML + """<div><table style="width:100%">
                        <thead><tr><td><b style="font-family:'Open Sans';font-size:13px">Action</b></td><td><b style="font-family:'Open Sans';font-size:13px">Instance Type</b></td><td><b style="font-family:'Open Sans';font-size:13px">Region</b></td><td><b style="font-family:'Open Sans';font-size:13px">Cache Engine</b></td><td><b style="font-family:'Open Sans';font-size:13px">Current Generation</b></td><td><b style="font-family:'Open Sans';font-size:13px">Upfront Cost</b></td><td><b style="font-family:'Open Sans';font-size:13px">Estimated Monthly Savings</b></td></tr></thead>
                        <tbody>"""
                        ri_details = ri_results['Recommendations'][0]['RecommendationDetails']
                        max_impact = {}
                        index = 0
                        
                        for ri in ri_details:
                            max_impact[index] = decimal.Decimal(ri['EstimatedMonthlySavingsAmount'])
                            index += 1
                        sort_max = sorted(max_impact.items(), key=lambda x:x[1], reverse=True)
                        sort_max = dict(sort_max)

                        # writing to excel
                        owner_name = aws_accounts[account][0]['acc_owner'].split('@')[0]
                        ws = wb_ri[owner_name]
                        ws.append([])
                        bold_cell = Cell(ws, value='Amazon ElastiCache for account {}'.format(aws_accounts[account][0]['acc_name']))
                        ws.append([bold_cell])
                        ws.append(['Action', 'Instance Type', 'Region', 'Cache Engine', 'Current Generation', 'Upfront Cost', 'Estimated Monthly Savings'])

                        for i in sort_max:
                            detail = ri_details[i]
                            
                            action = "Buy {0} {1}".format(detail['RecommendedNumberOfInstancesToPurchase'],detail['InstanceDetails']['ElastiCacheInstanceDetails']['NodeType'])
                            instancetype = detail['InstanceDetails']['ElastiCacheInstanceDetails']['NodeType']
                            rgn = detail['InstanceDetails']['ElastiCacheInstanceDetails']['Region']
                            engine = detail['InstanceDetails']['ElastiCacheInstanceDetails']['ProductDescription']
                            current_gen = str(detail['InstanceDetails']['ElastiCacheInstanceDetails']['CurrentGeneration'])
                            upcost = decimal.Decimal(detail['UpfrontCost'])
                            upcost_rounded = str(upcost.quantize(decimal.Decimal('0.00')))
                            est_saving = decimal.Decimal(detail['EstimatedMonthlySavingsAmount'])
                            est_sav_rounded = str(est_saving.quantize(decimal.Decimal('0.00')))
                            BODY_HTML = BODY_HTML + "<tr><td>"+action+"</td><td>"+instancetype+"</td><td>"+rgn+"</td><td>"+engine+"</td><td>"+current_gen+"</td><td>"+upcost_rounded+"</td><td>"+est_sav_rounded+"</td></tr>"

                            # writing to excel
                            ws.append([
                                action,
                                instancetype,
                                rgn,
                                engine,
                                current_gen,
                                upcost_rounded,
                                est_sav_rounded
                            ])
                        BODY_HTML = BODY_HTML + "<tr height = 20px></tr>"
                        BODY_HTML = BODY_HTML + "</tbody></table></div>"
                        
                    elif riservice == 'Amazon OpenSearch Service':
                        BODY_HTML = BODY_HTML + "<h4>RI Recommendations for {0}:</h4>".format(riservice)
                        BODY_HTML = BODY_HTML + """<div><table style="width:100%">
                        <thead><tr><td><b style="font-family:'Open Sans';font-size:13px">Action</b></td><td><b style="font-family:'Open Sans';font-size:13px">Instance Type</b></td><td><b style="font-family:'Open Sans';font-size:13px">Region</b></td><td><b style="font-family:'Open Sans';font-size:13px">Current Generation</b></td><td><b style="font-family:'Open Sans';font-size:13px">Upfront Cost</b></td><td><b style="font-family:'Open Sans';font-size:13px">Estimated Monthly Savings</b></td></tr></thead>
                        <tbody>"""
                        ri_details = ri_results['Recommendations'][0]['RecommendationDetails']
                        max_impact = {}
                        index = 0
                        
                        for ri in ri_details:
                            max_impact[index] = decimal.Decimal(ri['EstimatedMonthlySavingsAmount'])
                            index += 1
                        sort_max = sorted(max_impact.items(), key=lambda x:x[1], reverse=True)
                        sort_max = dict(sort_max)

                        # writing to excel
                        owner_name = aws_accounts[account][0]['acc_owner'].split('@')[0]
                        ws = wb_ri[owner_name]
                        ws.append([])
                        bold_cell = Cell(ws, value='Amazon OpenSearch Service for account {}'.format(aws_accounts[account][0]['acc_name']))
                        ws.append([bold_cell])
                        ws.append(['Action', 'Instance Type', 'Region', 'Current Generation', 'Upfront Cost', 'Estimated Monthly Savings'])

                        for i in sort_max:
                            detail = ri_details[i]
                            
                            action = "Buy {0} {1}".format(detail['RecommendedNumberOfInstancesToPurchase'],detail['InstanceDetails']['ESInstanceDetails']['InstanceSize'])
                            instancetype = detail['InstanceDetails']['ESInstanceDetails']['InstanceSize']
                            rgn = detail['InstanceDetails']['ESInstanceDetails']['Region']
                            current_gen = str(detail['InstanceDetails']['ESInstanceDetails']['CurrentGeneration'])
                            upcost = decimal.Decimal(detail['UpfrontCost'])
                            upcost_rounded = str(upcost.quantize(decimal.Decimal('0.00')))
                            est_saving = decimal.Decimal(detail['EstimatedMonthlySavingsAmount'])
                            est_sav_rounded = str(est_saving.quantize(decimal.Decimal('0.00')))
                            BODY_HTML = BODY_HTML + "<tr><td>"+action+"</td><td>"+instancetype+"</td><td>"+rgn+"</td><td>"+current_gen+"</td><td>"+upcost_rounded+"</td><td>"+est_sav_rounded+"</td></tr>"

                            # writing to excel
                            ws.append([
                                action,
                                instancetype,
                                rgn,
                                current_gen,
                                upcost_rounded,
                                est_sav_rounded
                            ])
                        BODY_HTML = BODY_HTML + "<tr height = 20px></tr>"
                        BODY_HTML = BODY_HTML + "</tbody></table></div>"
                    
                    elif riservice == 'Amazon Elastic Compute Cloud - Compute':
                        BODY_HTML = BODY_HTML + "<h4>RI Recommendations for {0}:</h4>".format(riservice)
                        BODY_HTML = BODY_HTML + """<div><table style="width:100%">
                        <thead><tr><td><b style="font-family:'Open Sans';font-size:13px">Action</b></td><td><b style="font-family:'Open Sans';font-size:13px">Instance Type</b></td><td><b style="font-family:'Open Sans';font-size:13px">Platform</b></td><td><b style="font-family:'Open Sans';font-size:13px">Region</b></td><td><b style="font-family:'Open Sans';font-size:13px">Current Generation</b></td><td><b style="font-family:'Open Sans';font-size:13px">Upfront Cost</b></td><td><b style="font-family:'Open Sans';font-size:13px">Estimated Monthly Savings</b></td></tr></thead>
                        <tbody>"""
                        ri_details = ri_results['Recommendations'][0]['RecommendationDetails']
                        max_impact = {}
                        index = 0
                        
                        for ri in ri_details:
                            max_impact[index] = decimal.Decimal(ri['EstimatedMonthlySavingsAmount'])
                            index += 1
                        sort_max = sorted(max_impact.items(), key=lambda x:x[1], reverse=True)
                        sort_max = dict(sort_max)

                        # writing to excel
                        owner_name = aws_accounts[account][0]['acc_owner'].split('@')[0]
                        ws = wb_ri[owner_name]
                        ws.append([])
                        bold_cell = Cell(ws, value='Amazon EC2 for account {}'.format(aws_accounts[account][0]['acc_name']))
                        ws.append([bold_cell])
                        ws.append(['Action', 'Instance Type', 'Platform', 'Region', 'Current Generation', 'Upfront Cost', 'Estimated Monthly Savings'])

                        for i in sort_max:
                            detail = ri_details[i]
                            
                            action = "Buy {0} {1}".format(detail['RecommendedNumberOfInstancesToPurchase'],detail['InstanceDetails']['EC2InstanceDetails']['InstanceType'])
                            instancetype = detail['InstanceDetails']['EC2InstanceDetails']['InstanceType']
                            platform = detail['InstanceDetails']['EC2InstanceDetails']['Platform']
                            rgn = detail['InstanceDetails']['EC2InstanceDetails']['Region']
                            current_gen = str(detail['InstanceDetails']['EC2InstanceDetails']['CurrentGeneration'])
                            upcost = decimal.Decimal(detail['UpfrontCost'])
                            upcost_rounded = str(upcost.quantize(decimal.Decimal('0.00')))
                            est_saving = decimal.Decimal(detail['EstimatedMonthlySavingsAmount'])
                            est_sav_rounded = str(est_saving.quantize(decimal.Decimal('0.00')))
                            BODY_HTML = BODY_HTML + "<tr><td>"+action+"</td><td>"+instancetype+"</td><td>"+platform+"</td><td>"+rgn+"</td><td>"+current_gen+"</td><td>"+upcost_rounded+"</td><td>"+est_sav_rounded+"</td></tr>"

                            # writing to excel
                            ws.append([
                                action,
                                instancetype,
                                platform,
                                rgn,
                                current_gen,
                                upcost_rounded,
                                est_sav_rounded
                            ])
                        BODY_HTML = BODY_HTML + "<tr height = 20px></tr>"
                        BODY_HTML = BODY_HTML + "</tbody></table></div>"
                    elif riservice == 'Amazon Redshift':
                        BODY_HTML = BODY_HTML + "<h4>RI Recommendations for {0}:</h4>".format(riservice)
                        BODY_HTML = BODY_HTML + """<div><table style="width:100%">
                        <thead><tr><td><b style="font-family:'Open Sans';font-size:13px">Action</b></td><td><b style="font-family:'Open Sans';font-size:13px">Instance Type</b></td><td><b style="font-family:'Open Sans';font-size:13px">Region</b></td><td><b style="font-family:'Open Sans';font-size:13px">SizeFlex Eligible</b></td><td><b style="font-family:'Open Sans';font-size:13px">Current Generation</b></td><td><b style="font-family:'Open Sans';font-size:13px">Upfront Cost</b></td><td><b style="font-family:'Open Sans';font-size:13px">Estimated Monthly Savings</b></td></tr></thead>
                        <tbody>"""
                        ri_details = ri_results['Recommendations'][0]['RecommendationDetails']
                        max_impact = {}
                        index = 0
                        
                        for ri in ri_details:
                            max_impact[index] = decimal.Decimal(ri['EstimatedMonthlySavingsAmount'])
                            index += 1
                        sort_max = sorted(max_impact.items(), key=lambda x:x[1], reverse=True)
                        sort_max = dict(sort_max)

                        # writing to excel
                        owner_name = aws_accounts[account][0]['acc_owner'].split('@')[0]
                        ws = wb_ri[owner_name]
                        ws.append([])
                        bold_cell = Cell(ws, value='Amazon Redshift for account {}'.format(aws_accounts[account][0]['acc_name']))
                        ws.append([bold_cell])
                        ws.append(['Action', 'Instance Type', 'Region', 'SizeFlex Eligible', 'Current Generation', 'Upfront Cost', 'Estimated Monthly Savings'])

                        for i in sort_max:
                            detail = ri_details[i]
                            
                            action = "Buy {0} {1}".format(detail['RecommendedNumberOfInstancesToPurchase'],detail['InstanceDetails']['RedshiftInstanceDetails']['NodeType'])
                            instancetype = detail['InstanceDetails']['RedshiftInstanceDetails']['NodeType']
                            rgn = detail['InstanceDetails']['RedshiftInstanceDetails']['Region']
                            sizeflex = str(detail['InstanceDetails']['RedshiftInstanceDetails']['SizeFlexEligible'])
                            current_gen = str(detail['InstanceDetails']['RedshiftInstanceDetails']['CurrentGeneration'])
                            upcost = decimal.Decimal(detail['UpfrontCost'])
                            upcost_rounded = str(upcost.quantize(decimal.Decimal('0.00')))
                            est_saving = decimal.Decimal(detail['EstimatedMonthlySavingsAmount'])
                            est_sav_rounded = str(est_saving.quantize(decimal.Decimal('0.00')))
                            BODY_HTML = BODY_HTML + "<tr><td>"+action+"</td><td>"+instancetype+"</td><td>"+rgn+"</td><td>"+sizeflex+"</td><td>"+current_gen+"</td><td>"+upcost_rounded+"</td><td>"+est_sav_rounded+"</td></tr>"

                            # writing to excel
                            ws.append([
                                action,
                                instancetype,
                                rgn,
                                sizeflex,
                                current_gen,
                                upcost_rounded,
                                est_sav_rounded
                            ])
                        BODY_HTML = BODY_HTML + "<tr height = 20px></tr>"
                        BODY_HTML = BODY_HTML + "</tbody></table></div>"
            BODY_HTML = BODY_HTML + "<h4>######################################</h4>"
    
    # setting excel file column width
    for sheet in wb.sheetnames: 
        ws = wb[sheet]
        col_width = max(len(str(cell.value)) for row in ws for cell in row)
        for col_idx, column in enumerate(ws.columns, start=1):
            column_letter = get_column_letter(col_idx)
            adjusted_width = (col_width + 2) * 0.8
            ws.column_dimensions[column_letter].width = adjusted_width
    
    for sheet in wb_ri.sheetnames: 
        ws = wb_ri[sheet]
        col_width = max(len(str(cell.value)) for row in ws for cell in row)
        for col_idx, column in enumerate(ws.columns, start=1):
            column_letter = get_column_letter(col_idx)
            adjusted_width = (col_width + 2) * 0.8
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save('/tmp/AWSCOST_WeeklyReport_{0}_to_{1}.xlsx'.format(start_week,end_week_title))
    wb_ri.save('/tmp/AWSCOST_RIReport_{0}_to_{1}.xlsx'.format(start_week,end_week_title))
    xlfile = ('/tmp/AWSCOST_WeeklyReport_{0}_to_{1}.xlsx'.format(start_week,end_week_title))
    xlrifile = ('/tmp/AWSCOST_RIReport_{0}_to_{1}.xlsx'.format(start_week,end_week_title))

    tmp_files = list()
    tmp_files = [xlfile, xlrifile]

    BODY_HTML = BODY_HTML + """Thank you,<br>
    DevOps Team<br>
    DevOps@company.com"""

    upload_file(xlfile, 'aws-ce-reports')
    send_email_with_attachment(start_week, end_week_title, BODY_HTML, tmp_files, SES_REGION)